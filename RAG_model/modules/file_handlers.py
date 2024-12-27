import os
from typing import List, Union
from PyPDF2 import PdfReader
import openpyxl
from pathlib import Path
import magic

class DocumentProcessor:
    def __init__(self):
        self.supported_extensions = {'.pdf', '.txt', '.xlsx', '.docx'}

    def process_file(self, file_path: Union[str, Path]) -> List[str]:
        """Process a file and return a list of text chunks."""
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        mime = magic.Magic(mime=True)
        file_type = mime.from_file(str(file_path))

        if file_type.startswith('text/'):
            return self._process_text_file(file_path)
        elif file_type == 'application/pdf':
            return self._process_pdf(file_path)
        elif file_type.startswith('application/vnd.openxmlformats-officedocument.spreadsheetml'):
            return self._process_excel(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

    def _process_text_file(self, file_path: Path) -> List[str]:
        """Process a text file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                text = file.read()
            return self._chunk_text(text)
        except Exception as e:
            raise Exception(f"Error processing text file: {str(e)}")

    def _process_pdf(self, file_path: Path) -> List[str]:
        """Process a PDF file."""
        try:
            reader = PdfReader(str(file_path))
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n"
            return self._chunk_text(text)
        except Exception as e:
            raise Exception(f"Error processing PDF file: {str(e)}")

    def _process_excel(self, file_path: Path) -> List[str]:
        """Process an Excel file."""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            texts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                sheet_text = f"Sheet: {sheet}\n"
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) for cell in row if cell is not None)
                    if row_text:
                        sheet_text += row_text + "\n"
                texts.extend(self._chunk_text(sheet_text))
            return texts
        except Exception as e:
            raise Exception(f"Error processing Excel file: {str(e)}")

    def _chunk_text(self, text: str, chunk_size: int = 1000) -> List[str]:
        """Split text into chunks of approximately equal size."""
        chunks = []
        current_chunk = ""
        
        # Split by sentences (simple approach)
        sentences = text.replace('\n', ' ').split('.')
        
        for sentence in sentences:
            sentence = sentence.strip() + '.'
            if len(current_chunk) + len(sentence) > chunk_size and current_chunk:
                chunks.append(current_chunk.strip())
                current_chunk = sentence
            else:
                current_chunk += ' ' + sentence
                
        if current_chunk:
            chunks.append(current_chunk.strip())
            
        return chunks
